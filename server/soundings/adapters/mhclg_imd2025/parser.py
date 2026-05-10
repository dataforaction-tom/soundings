"""Pure parser for the IMD 2025 Excel workbook.

Headers are matched case-insensitively against the strings pinned in
ADR-0002. Returns one ImdRow per (LSOA, indicator); aggregation to
LTLA is handled separately by the loader.
"""

import io
from dataclasses import dataclass

from openpyxl import load_workbook

# (Soundings indicator key, header substring to match in the sheet)
COLUMN_TO_INDICATOR: list[tuple[str, str]] = [
    ("deprivation.imd.score", "index of multiple deprivation (imd) score"),
    ("deprivation.imd.decile", "index of multiple deprivation (imd) decile"),
    ("deprivation.imd.income_score", "income score"),
    ("deprivation.imd.employment_score", "employment score"),
    ("deprivation.imd.health_score", "health deprivation"),
    ("deprivation.imd.education_score", "education, skills"),
    ("deprivation.idaci", "income deprivation affecting children"),
    ("deprivation.idaopi", "income deprivation affecting older people"),
]


@dataclass(frozen=True)
class ImdRow:
    lsoa_code: str
    indicator_key: str
    value: float


def parse_imd_xlsx(blob: bytes) -> list[ImdRow]:
    wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        return []

    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if headers is None:
        return []
    header_index: dict[int, str] = {
        idx: str(h).strip().lower() for idx, h in enumerate(headers) if h is not None
    }

    # Find the LSOA code column.
    lsoa_col = next(
        (idx for idx, h in header_index.items() if "lsoa code" in h),
        None,
    )
    if lsoa_col is None:
        return []

    # Map each indicator key to the matching column.
    indicator_cols: dict[str, int] = {}
    for indicator_key, header_substr in COLUMN_TO_INDICATOR:
        for idx, h in header_index.items():
            if header_substr in h:
                indicator_cols[indicator_key] = idx
                break

    out: list[ImdRow] = []
    for row in rows_iter:
        lsoa_code = row[lsoa_col]
        if not lsoa_code:
            continue
        for indicator_key, col_idx in indicator_cols.items():
            raw = row[col_idx]
            if raw is None:
                continue
            try:
                value = float(raw)
            except (TypeError, ValueError):
                continue
            out.append(
                ImdRow(
                    lsoa_code=str(lsoa_code).strip(),
                    indicator_key=indicator_key,
                    value=value,
                )
            )
    return out
